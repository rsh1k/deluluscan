"""tests.test_exporters — CSV / XLSX / JUnit output.

Three things these tests exist to prevent:

1. **A false positive appearing in a work queue.** A spreadsheet row reads as
   "fix this". A refuted candidate must never land in the findings rows.
2. **A formula executing when someone opens the report.** Finding titles carry
   attacker-supplied strings by construction; a title beginning "=" is a live
   formula in Excel and LibreOffice.
3. **CI failing on accepted behaviour.** An observation the product owner has
   accepted must not break a build, or the gate gets ignored.

Run: python3 -m tests.test_exporters
"""
from __future__ import annotations

import os
import sys
import tempfile
import xml.etree.ElementTree as ET

from deluluscan.reporting import exporters as ex

_checks = 0
_failures: list[str] = []


def check(cond: bool, label: str) -> None:
    global _checks
    _checks += 1
    if cond:
        print(f"PASS  {label}")
    else:
        _failures.append(label)
        print(f"FAIL  {label}")


def payload() -> dict:
    """A payload with one reported finding, one observation, one refuted."""
    return {
        "target": "http://127.0.0.1:8080",
        "date": "2026-08-23T00:00:00Z",
        "meta": {
            "report_include": {"ids": ["rep1"]},
            "scoring": {"system": "CVSS v3.1 Base"},
            "image": {"tag": "target/target:1.2.5"},
            "coverage": {"endpoints_discovered": 745, "endpoints_probed": 725},
        },
        "findings": [
            {"id": "rep1", "title": "Reported finding", "severity": "high",
             "vuln_class": "rate_limit", "endpoint": "POST /api/v1/authentication",
             "verdict": "true_positive", "exploitability": "exploitable",
             "confidence": "confirmed", "description": "a real one",
             "detail": {"report": {
                 "taxonomy": {"owasp_2025": "A02:2025", "owasp_api_top10": "API4",
                              "cwe": ["CWE-307"]},
                 "cvss": {"base_score": 7.5, "vector": "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:N/A:N"},
                 "compliance": {"frameworks": {
                     "PCI-DSS v4.0.1": [{"id": "8.3.4", "title": "t", "basis": "b"}],
                     "SOC 2 (2017 TSC)": [{"id": "CC6.1", "title": "t", "basis": "b"}],
                     "ISO/IEC 27001:2022": [{"id": "A.8.5", "title": "t", "basis": "b"}]}},
                 "observed": {"status_by_identity": {"anonymous": 200}, "requests_captured": 2},
             }}},
            {"id": "obs1", "title": "Accepted behaviour", "severity": "medium",
             "vuln_class": "info_leak", "endpoint": "GET /x",
             "verdict": "true_positive", "exploitability": "exploitable",
             "confidence": "confirmed", "description": "accepted",
             "detail": {"observation": True, "disposition": "accepted by product owner",
                        "report": {"taxonomy": {}, "observed": {}}}},
            {"id": "fp1", "title": "=cmd|'/c calc'!A1", "severity": "info",
             "vuln_class": "sqli", "endpoint": "", "verdict": "false_positive",
             "exploitability": "not_exploitable", "confidence": "confirmed",
             "description": "did not reproduce",
             "detail": {"refuted": True, "origin": "prior report"}},
        ],
    }


def test_csv_excludes_refuted_by_default():
    rows = ex.to_csv(payload()).splitlines()
    body = rows[1:]
    check(len(body) == 2, "CSV carries the reported finding and the observation, not the refuted one")
    check(all("did not reproduce" not in r for r in body),
          "a refuted candidate never appears in the default CSV work queue")
    check(any("Reported finding" in r for r in body), "the reported finding is present")


def test_csv_can_include_refuted_explicitly():
    rows = ex.to_csv(payload(), include_refuted=True).splitlines()[1:]
    check(len(rows) == 3, "include_refuted=True produces the full adjudication record")


def test_formula_injection_is_neutralised():
    """A title starting with '=' must not open as a live formula."""
    rows = ex.to_csv(payload(), include_refuted=True).splitlines()
    target = [r for r in rows if "calc" in r]
    check(bool(target), "the hostile title is present in the export")
    # csv quotes the field; the value inside must be prefixed with an apostrophe.
    check(all("\"'=cmd" in r or "'=cmd" in r for r in target),
          "a leading '=' is escaped so the cell is inert text, not a formula")
    for lead in ("=", "+", "-", "@"):
        check(ex._sanitize(f"{lead}danger").startswith("'"),
              f"a leading '{lead}' is neutralised")
    check(ex._sanitize("harmless") == "harmless", "benign text is left untouched")


def test_control_characters_stripped():
    check("\x00" not in ex._sanitize("a\x00b"), "NUL is stripped (Excel refuses it)")
    check(ex._sanitize("a\x07b") == "a b", "control characters become spaces")


def test_absent_cvss_is_blank_not_zero():
    """A zero would sort as 'safest' and read as a measured score."""
    row = ex.finding_row(payload()["findings"][1])
    check(row["cvss_score"] == "", "an unscored finding exports an empty score, not 0")
    check(row["cvss_vector"] == "", "an unscored finding exports an empty vector")


def test_compliance_columns_populate():
    row = ex.finding_row(payload()["findings"][0])
    check(row["pci_dss"] == "8.3.4", "PCI control id lands in its column")
    check(row["soc2"] == "CC6.1", "SOC 2 control id lands in its column")
    check(row["iso_27001"] == "A.8.5", "ISO control id lands in its column")


def test_severity_ordering():
    p = payload()
    ordered = ex.select_findings(p)
    check(ordered[0]["severity"] == "high", "higher severity sorts first")


def test_junit_only_fails_on_reported_findings():
    root = ET.fromstring(ex.to_junit(payload(), include_refuted=True))
    suite = root.find("testsuite")
    check(suite.get("failures") == "1",
          "only the reported finding fails the build")
    check(suite.get("skipped") == "2",
          "the observation and the refuted candidate are skipped, not failures")
    names = [tc.get("name") for tc in suite.iter("testcase")]
    check(any("observation" in (n or "") for n in names),
          "the observation still appears, so it stays visible in CI")


def test_junit_accepted_observation_never_breaks_the_build():
    """A payload with nothing reported must produce a green build."""
    p = payload()
    p["meta"]["report_include"]["ids"] = []
    p["findings"] = [f for f in p["findings"] if f["id"] != "rep1"]
    root = ET.fromstring(ex.to_junit(p, include_refuted=True))
    suite = root.find("testsuite")
    check(suite.get("failures") == "0",
          "accepted observations alone do not fail CI")


def test_junit_severity_gate_is_configurable():
    root = ET.fromstring(ex.to_junit(payload(), fail_on=("critical",)))
    check(root.find("testsuite").get("failures") == "0",
          "a high finding does not fail a critical-only gate")


def test_junit_is_well_formed_xml_with_hostile_content():
    xml = ex.to_junit(payload(), include_refuted=True)
    root = ET.fromstring(xml)          # raises if malformed
    check(root.tag == "testsuites", "JUnit output parses as XML despite hostile titles")
    check(xml.startswith("<?xml"), "JUnit output carries an XML declaration")


def test_xlsx_writes_and_separates_refuted():
    try:
        from openpyxl import load_workbook
    except ImportError:
        check(True, "openpyxl absent — XLSX test skipped")
        return
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "f.xlsx")
        ex.write_xlsx(payload(), path)
        check(os.path.getsize(path) > 0, "XLSX file is written")
        wb = load_workbook(path)
        check("Findings" in wb.sheetnames, "workbook has a Findings sheet")
        check("Refuted (false positives)" in wb.sheetnames,
              "refuted candidates get their own sheet, not findings rows")
        check("Scan metadata" in wb.sheetnames, "workbook records scan provenance")
        ws = wb["Findings"]
        titles = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        check(all("calc" not in str(t) for t in titles),
              "the refuted hostile row is not in the Findings sheet")
        check(ws.freeze_panes == "A2", "header row is frozen for triage")


def test_export_dispatch_rejects_unknown_format():
    try:
        ex.export(payload(), "docx", "/tmp/x")
        check(False, "unknown format raises")
    except ValueError as e:
        check("unknown export format" in str(e), "unknown format raises a clear ValueError")


def test_csv_roundtrips_through_reader():
    import csv as _csv
    import io as _io
    rows = list(_csv.DictReader(_io.StringIO(ex.to_csv(payload()))))
    check(len(rows) == 2, "CSV parses back to the expected row count")
    check(rows[0]["title"] == "Reported finding", "field values survive the round trip")


def main() -> int:
    print("== exporters ==")
    for fn in (test_csv_excludes_refuted_by_default,
               test_csv_can_include_refuted_explicitly,
               test_formula_injection_is_neutralised,
               test_control_characters_stripped,
               test_absent_cvss_is_blank_not_zero,
               test_compliance_columns_populate,
               test_severity_ordering,
               test_junit_only_fails_on_reported_findings,
               test_junit_accepted_observation_never_breaks_the_build,
               test_junit_severity_gate_is_configurable,
               test_junit_is_well_formed_xml_with_hostile_content,
               test_xlsx_writes_and_separates_refuted,
               test_export_dispatch_rejects_unknown_format,
               test_csv_roundtrips_through_reader):
        fn()
    print()
    if _failures:
        print(f"{len(_failures)} FAILED of {_checks} checks:")
        for f in _failures:
            print("  -", f)
        return 1
    print(f"{_checks}/{_checks} checks passed")
    return 0


if __name__ == "__main__":
    sys.exit(main())
