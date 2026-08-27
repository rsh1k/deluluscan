"""deluluscan.reporting.exporters — CSV, XLSX and JUnit output.

Deluluscan already emits HTML (the dashboard), JSON (results.json), SARIF, DOCX and
Markdown. Those cover reading a report and feeding a code-scanning platform.
Three consumers were not covered:

* **CSV** — triage in a spreadsheet, or import into a tracker. Plain text, no
  dependency, opens anywhere.
* **XLSX** — the same data with severity banding and frozen headers, for
  circulating to people who will filter and sort it.
* **JUnit XML** — CI gating. Almost every CI system renders JUnit natively, so
  a scan can fail a pipeline and show *which* finding failed it, without the
  platform needing to understand Deluluscan at all.

Two rules shared with the rest of the reporting layer:

1. **A refuted finding is never exported as a finding.** False positives are
   carried in a separate, clearly-labelled column/sheet or omitted entirely —
   never mixed into a row set someone will treat as a work queue.
2. **Evidence is never fabricated to fill a column.** An absent CVSS score is an
   empty cell, not a zero; zero would sort and filter as though it were real.
"""
from __future__ import annotations

import csv
import io
import re
import xml.etree.ElementTree as ET
from typing import Any, Iterable

# Column order is deliberate: identity first, then severity/score for sorting,
# then classification, then the evidence pointer.
COLUMNS = [
    "id", "title", "severity", "cvss_score", "cvss_vector", "vuln_class",
    "endpoint", "verdict", "exploitability", "confidence",
    "owasp_2025", "owasp_api_top10", "cwe",
    "pci_dss", "soc2", "iso_27001",
    "status_by_identity", "requests_captured", "description",
]

SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}

# Excel refuses control characters in cells, and a leading =/+/-/@ is executed
# as a formula by Excel and LibreOffice — a scanner that echoes attacker input
# into a report must not hand the reader a live formula.
_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")


def _sanitize(value: Any) -> str:
    """Render a value as spreadsheet-safe text.

    Neutralises CSV/XLSX formula injection by prefixing a single quote, and
    strips control characters that would make the file unopenable. This matters
    more here than in most exporters: finding titles and evidence contain
    attacker-supplied strings by construction.
    """
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        text = ", ".join(str(v) for v in value)
    elif isinstance(value, dict):
        text = ", ".join(f"{k}={v}" for k, v in value.items())
    else:
        text = str(value)
    text = _CONTROL_CHARS.sub(" ", text)
    if text.startswith(_FORMULA_LEAD):
        text = "'" + text
    return text


def _report(finding: dict) -> dict:
    return (finding.get("detail") or {}).get("report") or {}


def _compliance_ids(finding: dict, framework_substring: str) -> str:
    """Control IDs for one framework, as a comma-separated cell."""
    block = (_report(finding).get("compliance") or {}).get("frameworks") or {}
    for name, entries in block.items():
        if framework_substring.lower() in name.lower():
            return ", ".join(e.get("id", "") for e in entries)
    return ""


def finding_row(finding: dict) -> dict[str, str]:
    """One finding flattened to the export column set."""
    rep = _report(finding)
    tax = rep.get("taxonomy") or {}
    cvss = rep.get("cvss") or (finding.get("detail") or {}).get("cvss") or {}
    observed = rep.get("observed") or {}
    return {
        "id": _sanitize(finding.get("id")),
        "title": _sanitize(finding.get("title")),
        "severity": _sanitize(finding.get("severity")),
        # Absent score stays empty — a 0 would sort as "safest" and read as measured.
        "cvss_score": _sanitize(cvss.get("base_score")) if cvss else "",
        "cvss_vector": _sanitize(cvss.get("vector")) if cvss else "",
        "vuln_class": _sanitize(finding.get("vuln_class")),
        "endpoint": _sanitize(finding.get("endpoint")),
        "verdict": _sanitize(finding.get("verdict")),
        "exploitability": _sanitize(finding.get("exploitability")),
        "confidence": _sanitize(finding.get("confidence")),
        "owasp_2025": _sanitize(tax.get("owasp_2025")),
        "owasp_api_top10": _sanitize(tax.get("owasp_api_top10")),
        "cwe": _sanitize(tax.get("cwe")),
        "pci_dss": _sanitize(_compliance_ids(finding, "PCI")),
        "soc2": _sanitize(_compliance_ids(finding, "SOC")),
        "iso_27001": _sanitize(_compliance_ids(finding, "ISO")),
        "status_by_identity": _sanitize(observed.get("status_by_identity")),
        "requests_captured": _sanitize(observed.get("requests_captured")),
        "description": _sanitize(finding.get("description")),
    }


def select_findings(payload: dict, *, include_refuted: bool = False,
                    include_observations: bool = True) -> list[dict]:
    """The findings an export should carry, in severity order.

    Refuted candidates are excluded by default: a spreadsheet row reads as work
    to do, and a false positive is not work to do. Pass include_refuted=True to
    produce the full adjudication record instead.
    """
    out = []
    for f in payload.get("findings", []):
        detail = f.get("detail") or {}
        if detail.get("refuted") and not include_refuted:
            continue
        if detail.get("observation") and not include_observations:
            continue
        out.append(f)
    out.sort(key=lambda f: (SEVERITY_ORDER.get(f.get("severity", "info"), 9),
                            -(((_report(f).get("cvss")) or {}).get("base_score") or 0)))
    return out


# --------------------------------------------------------------------------
# CSV
# --------------------------------------------------------------------------
def to_csv(payload: dict, *, include_refuted: bool = False) -> str:
    """Findings as CSV text (UTF-8, RFC 4180 quoting)."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=COLUMNS, extrasaction="ignore",
                            lineterminator="\n")
    writer.writeheader()
    for f in select_findings(payload, include_refuted=include_refuted):
        writer.writerow(finding_row(f))
    return buf.getvalue()


def write_csv(payload: dict, path: str, *, include_refuted: bool = False) -> str:
    # utf-8-sig: Excel misreads plain UTF-8 as the local codepage and mangles
    # non-ASCII in titles; the BOM is what makes it open correctly.
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        fh.write(to_csv(payload, include_refuted=include_refuted))
    return path


# --------------------------------------------------------------------------
# XLSX
# --------------------------------------------------------------------------
_SEV_FILL = {
    "critical": "FFC7CE", "high": "FFD9B3", "medium": "FFEB9C",
    "low": "DDEBF7", "info": "EDEDED",
}


def write_xlsx(payload: dict, path: str, *, include_refuted: bool = True) -> str:
    """Findings as a formatted workbook.

    Refuted candidates go on their OWN sheet rather than into the findings rows,
    so the primary sheet stays a work queue while the adjudication record
    remains available in the same file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:                       # pragma: no cover
        raise RuntimeError(
            "XLSX export needs openpyxl: pip install openpyxl --break-system-packages"
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F172A")

    def write_sheet(sheet, rows, columns):
        sheet.append(columns)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for row in rows:
            sheet.append([row.get(c, "") for c in columns])
            sev = (row.get("severity") or "").lower()
            if sev in _SEV_FILL:
                sheet.cell(row=sheet.max_row, column=columns.index("severity") + 1) \
                     .fill = PatternFill("solid", fgColor=_SEV_FILL[sev])
        sheet.freeze_panes = "A2"
        if rows:
            sheet.auto_filter.ref = (
                f"A1:{get_column_letter(len(columns))}{len(rows) + 1}")
        for i, col in enumerate(columns, start=1):
            longest = max([len(col)] + [len(str(r.get(col, ""))) for r in rows] or [len(col)])
            sheet.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 60)

    findings = select_findings(payload, include_refuted=False)
    write_sheet(ws, [finding_row(f) for f in findings], COLUMNS)

    if include_refuted:
        refuted = [f for f in payload.get("findings", [])
                   if (f.get("detail") or {}).get("refuted")]
        if refuted:
            rs = wb.create_sheet("Refuted (false positives)")
            write_sheet(rs, [{"id": _sanitize(f.get("id")),
                              "title": _sanitize(f.get("title")),
                              "vuln_class": _sanitize(f.get("vuln_class")),
                              "origin": _sanitize((f.get("detail") or {}).get("origin")),
                              "why_not_a_finding": _sanitize(f.get("description"))}
                             for f in refuted],
                        ["id", "title", "vuln_class", "origin", "why_not_a_finding"])

    meta = payload.get("meta", {})
    ms = wb.create_sheet("Scan metadata")
    ms.append(["Property", "Value"])
    for cell in ms[1]:
        cell.font = header_font
        cell.fill = header_fill
    img = meta.get("image") or {}
    cov = meta.get("coverage") or {}
    for k, v in [("Target", payload.get("target")),
                 ("Scan date", payload.get("date")),
                 ("Image", img.get("tag")), ("Image digest", img.get("digest")),
                 ("Source commit", img.get("source_commit")),
                 ("Endpoints discovered", cov.get("endpoints_discovered")),
                 ("Endpoints probed", cov.get("endpoints_probed")),
                 ("Scoring system", (meta.get("scoring") or {}).get("system"))]:
        if v is not None:
            ms.append([k, _sanitize(v)])
    ms.column_dimensions["A"].width = 26
    ms.column_dimensions["B"].width = 80

    wb.save(path)
    return path


# --------------------------------------------------------------------------
# JUnit XML
# --------------------------------------------------------------------------
def to_junit(payload: dict, *, fail_on: Iterable[str] = ("critical", "high", "medium"),
             include_refuted: bool = False) -> str:
    """Findings as JUnit XML, for CI gating.

    Each finding is a testcase. A finding at or above the `fail_on` severities
    is a FAILURE (it breaks the build); anything below is recorded as a passing
    testcase so the count reflects what was assessed rather than only what hurt.

    Refuted candidates are emitted as SKIPPED when included — semantically
    right, and it keeps a CI dashboard from showing a false positive as a
    regression.
    """
    fail_set = {s.lower() for s in fail_on}
    all_findings = payload.get("findings", [])
    reported_ids = set((payload.get("meta", {}).get("report_include") or {}).get("ids") or [])

    # Only REPORTED findings can fail a build. An observation is behaviour the
    # product owner has accepted or that was shown not to be exploitable —
    # failing CI on it would train people to ignore the gate. Observations are
    # emitted as skipped so they stay visible without blocking.
    findings = [f for f in all_findings
                if f["id"] in reported_ids
                or (not reported_ids and not (f.get("detail") or {}).get("refuted")
                    and not (f.get("detail") or {}).get("observation"))]
    findings.sort(key=lambda f: (SEVERITY_ORDER.get(f.get("severity", "info"), 9),
                                 -(((_report(f).get("cvss")) or {}).get("base_score") or 0)))
    observations = [f for f in all_findings
                    if (f.get("detail") or {}).get("observation") and f["id"] not in reported_ids]
    refuted = ([f for f in all_findings
                if (f.get("detail") or {}).get("refuted")] if include_refuted else [])

    failures = sum(1 for f in findings if (f.get("severity") or "").lower() in fail_set)
    skipped_total = len(refuted) + len(observations)
    suite = ET.Element("testsuite", {
        "name": "deluluscan",
        "tests": str(len(findings) + skipped_total),
        "failures": str(failures),
        "errors": "0",
        "skipped": str(skipped_total),
    })
    target = payload.get("target") or ""
    props = ET.SubElement(suite, "properties")
    for key, value in (("target", target),
                       ("scan_date", payload.get("date") or ""),
                       ("scoring", (payload.get("meta", {}).get("scoring") or {}).get("system", ""))):
        if value:
            ET.SubElement(props, "property", {"name": key, "value": str(value)})

    for f in findings:
        rep = _report(f)
        cvss = rep.get("cvss") or {}
        classname = f"deluluscan.{f.get('vuln_class') or 'finding'}"
        case = ET.SubElement(suite, "testcase", {
            "classname": classname,
            "name": f"{f.get('title', 'finding')} [{f.get('endpoint') or 'n/a'}]",
        })
        if (f.get("severity") or "").lower() in fail_set:
            detail = [
                f"Severity : {f.get('severity')}",
                f"CVSS     : {cvss.get('base_score', 'n/a')} {cvss.get('vector', '')}".rstrip(),
                f"Endpoint : {f.get('endpoint') or 'n/a'}",
                f"Verdict  : {f.get('verdict')} / {f.get('exploitability')}",
                "",
                (f.get("description") or "").strip(),
            ]
            failure = ET.SubElement(case, "failure", {
                "message": f"{f.get('severity', '').upper()}: {f.get('title', '')}",
                "type": str(f.get("vuln_class") or "finding"),
            })
            failure.text = "\n".join(detail)

    for f in observations:
        case = ET.SubElement(suite, "testcase", {
            "classname": f"deluluscan.observation.{f.get('vuln_class') or 'other'}",
            "name": f"{f.get('title', 'observation')} (observation)",
        })
        dispo = (f.get("detail") or {}).get("disposition") or (
            "recorded as an observation; not counted as a vulnerability")
        skipped = ET.SubElement(case, "skipped", {
            "message": "observation — not a reported vulnerability"})
        skipped.text = dispo

    for f in refuted:
        case = ET.SubElement(suite, "testcase", {
            "classname": "deluluscan.refuted",
            "name": f"{f.get('title', 'candidate')} (refuted)",
        })
        skipped = ET.SubElement(case, "skipped", {
            "message": "refuted by live re-testing — not a finding"})
        skipped.text = (f.get("description") or "").strip()

    suites = ET.Element("testsuites", {
        "name": "deluluscan", "tests": suite.get("tests"),
        "failures": suite.get("failures"), "errors": "0",
    })
    suites.append(suite)
    return ('<?xml version="1.0" encoding="UTF-8"?>\n'
            + ET.tostring(suites, encoding="unicode"))


def write_junit(payload: dict, path: str, **kwargs) -> str:
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(to_junit(payload, **kwargs))
    return path


FORMATS = {"csv": write_csv, "xlsx": write_xlsx, "junit": write_junit}


def export(payload: dict, fmt: str, path: str, **kwargs) -> str:
    """Write `payload` to `path` in `fmt`. Raises ValueError on unknown format."""
    key = fmt.lower().strip()
    if key not in FORMATS:
        raise ValueError(f"unknown export format {fmt!r}; known: {', '.join(sorted(FORMATS))}")
    return FORMATS[key](payload, path, **kwargs)
